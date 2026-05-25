"""
Export Service
Handles exporting scraped emails to various formats
"""

import csv
import io
import json
from typing import List, Dict, Optional
from datetime import datetime
from datetime import datetime

from app.models.user import ScrapedEmail


class ExportService:
    """Service for exporting email data to various formats"""

    # Column mappings for different CRMs
    CRM_MAPPINGS = {
        'default': {
            'email': 'email',
            'first_name': 'first_name',
            'last_name': 'last_name',
            'platform': 'source_platform',
            'source_username': 'source_username',
            'source_url': 'source_url',
            'bio': 'bio',
            'follower_count': 'follower_count',
            'verification_status': 'verification_status',
        },
        'instantly': {
            'email': 'email_address',
            'first_name': 'first_name',
            'last_name': 'last_name',
            'platform': 'source',
            'source_username': 'username',
        },
        'smartlead': {
            'email': 'email',
            'first_name': 'first_name',
            'last_name': 'last_name',
            'platform': 'platform',
        },
        'lemlist': {
            'email': 'email',
            'first_name': 'firstName',
            'last_name': 'lastName',
            'platform': '追道',
        },
        'hubspot': {
            'email': 'email',
            'first_name': 'firstname',
            'last_name': 'lastname',
        },
    }

    def __init__(self, crm_type: str = 'default'):
        self.crm_mapping = self.CRM_MAPPINGS.get(crm_type, self.CRM_MAPPINGS['default'])

    def export_to_csv(self, emails: List[ScrapedEmail], 
                      include_unverified: bool = False) -> str:
        """Export emails to CSV format"""
        output = io.StringIO()
        
        # Get columns based on CRM mapping
        columns = list(self.crm_mapping.keys())
        
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        
        for email in emails:
            if not include_unverified and email.verification_status.value not in ['verified', 'risky']:
                continue
            
            row = {
                'email': email.email,
                'first_name': email.first_name or '',
                'last_name': email.last_name or '',
                'platform': email.platform.value,
                'source_username': email.source_username or '',
                'source_url': email.source_url or '',
                'source_profile_url': email.source_profile_url or '',
                'bio': (email.bio or '')[:500] if email.bio else '',  # Limit bio length
                'follower_count': email.follower_count or 0,
                'verification_status': email.verification_status.value,
            }
            writer.writerow(row)
        
        return output.getvalue()

    def export_to_excel(self, emails: List[ScrapedEmail],
                        include_unverified: bool = False) -> bytes:
        """Export emails to Excel format"""
        data = []
        
        for email in emails:
            if not include_unverified and email.verification_status.value not in ['verified', 'risky']:
                continue
            
            row = {
                'Email': email.email,
                'First Name': email.first_name or '',
                'Last Name': email.last_name or '',
                'Platform': email.platform.value,
                'Source Username': email.source_username or '',
                'Source URL': email.source_url or '',
                'Profile URL': email.source_profile_url or '',
                'Bio': (email.bio or '')[:500] if email.bio else '',
                'Follower Count': email.follower_count or 0,
                'Verification Status': email.verification_status.value,
                'Scraped Date': email.created_at.strftime('%Y-%m-%d %H:%M:%S') if email.created_at else '',
            }
            data.append(row)
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"

        if not data:
            # Handle empty data case gracefully
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Write rows
        for row in data:
            ws.append([row.get(h, '') for h in headers])

        # Auto-adjust column widths
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in col_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)

    def export_to_json(self, emails: List[ScrapedEmail],
                       include_unverified: bool = False) -> str:
        """Export emails to JSON format"""
        data = []
        
        for email in emails:
            if not include_unverified and email.verification_status.value not in ['verified', 'risky']:
                continue
            
            # Parse verification details if stored as string
            verification_details = {}
            if email.verification_details:
                try:
                    import ast
                    verification_details = ast.literal_eval(email.verification_details)
                except:
                    verification_details = {'raw': email.verification_details}
            
            row = {
                'email': email.email,
                'first_name': email.first_name,
                'last_name': email.last_name,
                'platform': email.platform.value,
                'source_username': email.source_username,
                'source_url': email.source_url,
                'source_profile_url': email.source_profile_url,
                'bio': email.bio,
                'follower_count': email.follower_count,
                'verification_status': email.verification_status.value,
                'verification_details': verification_details,
                'created_at': email.created_at.isoformat() if email.created_at else None,
            }
            data.append(row)
        
        return json.dumps({
            'export_date': datetime.utcnow().isoformat(),
            'total_records': len(data),
            'emails': data
        }, indent=2)

    def export_to_google_sheets_format(self, emails: List[ScrapedEmail]) -> List[List]:
        """Export for Google Sheets import"""
        # Header row
        headers = ['Email', 'First Name', 'Last Name', 'Platform', 
                   'Source Username', 'Profile URL', 'Follower Count', 'Verification Status']
        
        rows = [headers]
        
        for email in emails:
            if email.verification_status.value not in ['verified', 'risky']:
                continue
            
            row = [
                email.email,
                email.first_name or '',
                email.last_name or '',
                email.platform.value,
                email.source_username or '',
                email.source_profile_url or '',
                email.follower_count or 0,
                email.verification_status.value,
            ]
            rows.append(row)
        
        return rows

    @staticmethod
    def save_export_history(db, user_id: int, campaign_id: int, 
                            filename: str, format: str, record_count: int):
        """Save export history record"""
        from app.models.user import ExportHistory
        
        export = ExportHistory(
            user_id=user_id,
            campaign_id=campaign_id,
            filename=filename,
            format=format,
            record_count=record_count,
        )
        db.add(export)
        db.commit()
        
        return export


def generate_filename(campaign_name: str, format: str) -> str:
    """Generate a descriptive filename for exports"""
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    
    # Clean campaign name
    clean_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in campaign_name)
    clean_name = clean_name.strip().replace(' ', '_')[:30]
    
    return f"SocialScravio_{clean_name}_{timestamp}.{format}"